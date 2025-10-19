import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
import time
import xlwings as xw

excel_path = https://github.com/vikramt07/PIP-SC_team_dimension/blob/main/PIP-SC_team_dimension%20-%20Copy.xlsx

st.title("PIP-SC Team Dimension Calculation— USER_INPUT")

# --- Editable USER_INPUT setup ---
editable_columns = ["Country", "Project Code", "Scope", "Project Volume", "Project Duration"]
dropdown_columns = ["Country", "Project Code", "Scope"]
merge_titles = ["SRAN", "5G", "MW", "CW", "COMMON TEAM"]

# --- Load dropdown options from Sheet4 ---
sheet4_df = pd.read_excel(excel_path, sheet_name="Sheet4")
dropdown_options = {col: sheet4_df[col].dropna().unique().tolist() for col in dropdown_columns}

# --- Initialize session state ---
if "submitted" not in st.session_state:
    st.session_state.submitted = False

# --- Load USER_INPUT table ---
def load_user_table():
    wb = load_workbook(excel_path, data_only=False)
    ws = wb["Sheet1"]
    df_user = None
    min_col = min_row = 0
    for t in ws.tables:
        if t == "USER_INPUT":
            tbl_obj = ws.tables[t]
            min_col, min_row, max_col, max_row = range_boundaries(tbl_obj.ref)
            usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
            nrows = max_row - min_row + 1
            df_user = pd.read_excel(
                excel_path,
                sheet_name=ws.title,
                header=0,
                usecols=usecols,
                skiprows=min_row - 1,
                nrows=nrows
            )
            df_user.columns = [str(c).strip() for c in df_user.columns]
            df_user = df_user.loc[:, ~df_user.columns.duplicated(keep="first")]
            df_user = df_user.fillna("")
            # Convert numeric to whole numbers
            for col in df_user.columns:
                df_user[col] = df_user[col].map(lambda x: int(round(float(x))) if str(x).replace('.', '', 1).isdigit() else x)
            break
    return df_user, ws, min_col, min_row

# --- Load other tables (read-only) ---
def load_other_tables():
    wb = load_workbook(excel_path, data_only=False)
    all_tables = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for t in ws.tables:
            if t != "USER_INPUT":
                tbl_obj = ws.tables[t]
                ref = getattr(tbl_obj, "ref", None)
                if not ref:
                    continue
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
                nrows = max_row - (min_row - 1)
                df = pd.read_excel(
                    excel_path,
                    sheet_name=sheet_name,
                    header=0,
                    usecols=usecols,
                    skiprows=min_row - 1,
                    nrows=nrows
                )
                df.columns = [str(c).strip() for c in df.columns]
                df = df.fillna("")
                # Convert numeric to whole numbers
                for col in df.columns:
                    df[col] = df[col].map(lambda x: int(round(float(x))) if str(x).replace('.', '', 1).isdigit() else x)
                all_tables[t] = df
    return all_tables

# --- Render USER_INPUT ---
df_user, ws_user, min_col, min_row = load_user_table()
if df_user is None:
    st.warning("USER_INPUT table not found.")
else:
    column_config_dict = {}
    for col in df_user.columns:
        if col in dropdown_columns:
            column_config_dict[col] = st.column_config.SelectboxColumn(label=col, options=dropdown_options[col])
        elif col in ["Project Volume", "Project Duration"]:
            column_config_dict[col] = st.column_config.NumberColumn(label=col, step=1)
        else:
            column_config_dict[col] = st.column_config.Column(label=col, disabled=True)

    edited_df = st.data_editor(df_user, column_config=column_config_dict, width="stretch")

    if st.button("Submit USER_INPUT"):
        changed = False
        wb = load_workbook(excel_path, data_only=False)  # keep formulas intact
        ws = wb["Sheet1"]

        # Write only edited cells in USER_INPUT
        for row_idx in range(len(df_user)):
            for col_name in editable_columns:
                if col_name in df_user.columns:
                    old_val = df_user.at[row_idx, col_name]
                    new_val = edited_df.at[row_idx, col_name]
                    if old_val != new_val:
                        excel_col_idx = df_user.columns.get_loc(col_name) + min_col
                        excel_row_idx = min_row + row_idx + 1
                        ws.cell(row=excel_row_idx, column=excel_col_idx, value=new_val)
                        changed = True

        if changed:
            # 1️⃣ Save user edits using openpyxl
            wb.save(excel_path)
            wb.close()  # Close to release file lock

            # 2️⃣ Recalculate formulas using xlwings
            try:
                xw_app = xw.App(visible=False)
                xw_book = xw_app.books.open(excel_path)
                xw_book.app.calculate()
                xw_book.save()
                xw_book.close()
                xw_app.quit()
                st.success("✅ USER_INPUT saved and recalculated successfully!")
            except Exception as e:
                st.warning(f"Formula recalculation skipped: {e}")
        else:
            st.info("ℹ️ No changes detected.")


        # Set flag to render other tables
        st.session_state.submitted = True
        st.rerun()  # Stop current run; app reloads automatically

# --- Render other tables only after first submission ---
if st.session_state.submitted:
    other_tables = load_other_tables()
    for name, df in other_tables.items():
        if name.upper() == "TEAM_DIMENSION":
            st.subheader("TEAM_DIMENSION Table")
            num_cols = len(df.columns)
            html_table = "<table style='width:100%; border-collapse:collapse; text-align:center;' border='1'>"
            for _, row in df.iterrows():
                first_cell = str(row[df.columns[0]]).strip()
                if first_cell in merge_titles:
                    html_table += f"""
                        <tr style='background-color:#0E1117; color:white; font-weight:bold;'>
                            <td colspan='{num_cols}' style='text-align:center; font-size:15px; padding:6px;'>
                                {first_cell}
                            </td>
                        </tr>
                    """
                else:
                    html_table += "<tr>"
                    for col in df.columns:
                        html_table += f"<td>{row[col]}</td>"
                    html_table += "</tr>"
            html_table += "</table>"
            st.markdown(html_table, unsafe_allow_html=True)
        else:
            st.subheader(f"{name} Table")
            st.dataframe(df, width="stretch")


